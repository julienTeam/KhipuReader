"""
Khipu Translator — Multi-level translation engine.

Produces three levels of translation:
  Level 1 (Cord): Raw cord-by-cord data (type, value, syllabic reading, color, level)
  Level 2 (Record): Structured records grouped by cluster (spreadsheet view)
  Level 3 (Document): Interpreted reading with document-type detection

Reference: Sivan, J. (2026). "The Khipu as a Layered Information System."
"""

from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass, field
from typing import Optional

import pandas as pd

from khipu_translator.database import KhipuDB, KhipuRecord
from khipu_translator.dictionary import (
    DICTIONARY,
    DOMAIN_GLOSSARIES,
    GLOSSARY,
    analyze_morphology,
)
from khipu_translator.locke import cord_type, decode_locke_value
from khipu_translator.syllabary import (
    FIGURE_EIGHT_SYLLABLE,
    TURNS_TO_ONSET,
    TURNS_TO_SYLLABLE,
    VALID_TURNS,
)


# --- Color roles (validated cross-corpus, chi2=480, p<1e-17) -----------------

COLOR_ROLES: dict[str, str] = {
    "LK": "text/governance",
    "FB": "action",
    "PR": "action/material",
    "GG": "identity/justice",
    "DB": "objects",
    "LC": "ritual",
    "LB": "interrogative",
    "BG": "identity",
    "AB": "subject/mixed",
    "W":  "data/neutral",
    "MB": "category/data",
    "KB": "data",
    "B":  "geographic",
    "HB": "data",
}


# --- Data structures ---------------------------------------------------------


@dataclass
class CordTranslation:
    """Level 1: Translation of a single cord."""
    cord_id: int
    cord_type: str           # 'INT', 'STRING', 'EMPTY'
    level: int               # hierarchical depth (1 = pendant, 2 = subsidiary, ...)
    ordinal: int             # position within cluster (CLUSTER_ORDINAL)
    global_ordinal: int      # position on primary cord (CORD_ORDINAL)
    color: str               # OKR color code
    color_role: str          # semantic role of this color

    # Numerical channel (Locke)
    locke_value: Optional[int]
    locke_confidence: str    # 'exact', 'uniform_kvt', 'ambiguous', 'none'

    # Textual channel (ALBA)
    alba_reading: Optional[str]       # proposed syllabic reading
    alba_gloss_en: Optional[str]      # English gloss
    alba_gloss_fr: Optional[str]      # French gloss
    alba_confirmed: bool              # matches dictionary
    alba_confidence: str              # 'high', 'medium', 'low', 'unconfirmed', 'none'
    alba_compound: Optional[str]      # compound decomposition if applicable

    # Knot sequence: raw description of knots on this cord
    # e.g. "S10 S10 L3" or "L3 L4 E" or "S100 S10 L5"
    knot_sequence: str = ""

    # Numerical prefix (S-knots on STRING cords)
    s_prefix: int = 0

    # L-1 flag (NULL turns in OKR -> unreadable)
    is_l1_null: bool = False

    # Tree structure
    parent_cord_id: Optional[int] = None
    children: list[CordTranslation] = field(default_factory=list)

    def to_dict(self) -> dict:
        """Convert to a flat dictionary for export."""
        d = {
            "cord_id": self.cord_id,
            "type": self.cord_type,
            "level": self.level,
            "ordinal": self.ordinal,
            "global_ordinal": self.global_ordinal,
            "color": self.color,
            "color_role": self.color_role,
            "locke_value": self.locke_value,
            "locke_confidence": self.locke_confidence,
            "alba_reading": self.alba_reading,
            "alba_gloss_en": self.alba_gloss_en,
            "alba_gloss_fr": self.alba_gloss_fr,
            "alba_confirmed": self.alba_confirmed,
            "alba_confidence": self.alba_confidence,
            "alba_compound": self.alba_compound,
            "knot_sequence": self.knot_sequence,
            "s_prefix": self.s_prefix,
            "is_l1_null": self.is_l1_null,
            "parent_cord_id": self.parent_cord_id,
        }
        if self.children:
            d["children"] = [c.to_dict() for c in self.children]
        return d


@dataclass
class ClusterRecord:
    """Level 2: A single cluster (row in the khipu spreadsheet)."""
    cluster_index: int
    cluster_id: Optional[int]
    cords: list[CordTranslation]

    @property
    def total_value(self) -> int:
        return sum(c.locke_value or 0 for c in self.cords)

    @property
    def string_readings(self) -> list[str]:
        return [c.alba_reading for c in self.cords if c.alba_reading]

    @property
    def colors(self) -> list[str]:
        return [c.color for c in self.cords]

    def to_dict(self) -> dict:
        return {
            "cluster_index": self.cluster_index,
            "total_value": self.total_value,
            "string_readings": self.string_readings,
            "colors": self.colors,
            "cords": [c.to_dict() for c in self.cords],
        }


# --- Document-type detection -------------------------------------------------

DOCUMENT_PROFILES: dict[str, set[str]] = {
    # §4.1 Tribute register (detected via Calculator schema + high values)
    "tribute_register":        {"pataka", "kama", "kamay", "ypa", "llaqa"},
    # §4.2 Paired admin register (detected via paired khipu structure)
    "paired_register":         {"kama", "kamay", "ypa", "pataka", "qapa", "kapa",
                                 "llaqa", "pata"},
    # §4.3 Cadastral survey
    "cadastral_survey":        {"qaqa", "kaqa", "taqa", "chiqa", "paqa", "naqa",
                                 "siqa", "piqa", "qata", "qama"},
    # §4.4 Regional census (3 colors AB/MB/HB)
    "regional_census":         {"tama", "llaqa", "pataka", "kama", "pata"},
    # §4.5 Judicial dossier
    "judicial_proceeding":     {"llalla", "mana", "taq", "pay", "kay", "sina",
                                 "pi", "piqa", "taka", "maqa", "kata"},
    # §4.6 Formal admin text
    "formal_admin":            {"kama", "kamay", "ypa", "qapa", "kapa", "tayta",
                                 "panaka"},
    # §4.7 Identity/naming record
    "identity_naming":         {"kiki", "taki", "mama", "tata", "chiki"},
    # §4.8 Astronomical catalog (CORE words only — 2 confirmed: UR006, UR1145)
    "astronomical_journal":    {"kama", "mama", "paka", "maqa"},
    # §4.9 Mit'a labor register
    "labor_tribute":           {"maki", "taka", "kata", "paka", "maka", "paki",
                                 "makay", "takay", "kiki", "kaki", "tama"},
    # §4.10 Security register — MERGED INTO judicial_proceeding (too much overlap)
    # "security_register":       {"maka", "taka", "kata", "kama"},
    # §4.11 Interrogation form (flat + piqa)
    "interrogation_form":      {"piqa", "pi", "pita", "pika", "may"},
    # §4.12 Production dossier — MERGED INTO labor_tribute (too much overlap with cadastral)
    # "production_dossier":      {"maki", "paki", "tama", "kaki", "qaqa"},
    # §4.13 Oracle/ritual text
    "ritual_oracle":           {"taki", "naqa", "pama", "napa", "naku", "tina",
                                 "llapa", "nay", "waka"},
    # §4.14 Pre-Inca binary (detected via Form schema)
    "pre_inca_binary":         set(),  # detected by schema, not vocabulary
    # Agro-pastoral (from this session's reclassification)
    "agro_pastoral":           {"kaki", "kaqa", "wama", "chaqa", "paki", "taki",
                                 "taqa", "siqa", "maki"},
    # Kinship lineage
    "kinship_lineage":         {"kaka", "mama", "papa", "tata", "pana", "tayta",
                                 "panaka", "tayka", "nana"},
}


def detect_document_type(
    vocabulary: Counter,
    string_ratio: float = 0.0,
    color_counts: Optional[dict[str, int]] = None,
    total_cords: int = 0,
    string_cords: int = 0,
    num_clusters: int = 0,
    cluster_regularity: float = 0.0,
    sparsity: float = 0.0,
    n_colors: int = 1,
) -> tuple[str, float]:
    """
    Detect the most likely document type from vocabulary, structure, and
    statistical signals.

    Categories:
      - astronomical_journal:  3+ core astro words (kama/mama/paka/maqa + kaki)
      - agro_pastoral:         regular full table + agro vocab (kaqa, wama, chaqa)
      - cadastral_survey:      toponym vocabulary (-qa words)
      - judicial_proceeding:   action + kinship co-occurrence
      - labor_tribute:         maki + companions (kiki, kaki, tama)
      - ritual_oracle:         taki, pama, napa, nay
      - kinship_lineage:       pure kinship vocabulary
      - event_register:        >50% sparse, punctual data
      - regular_table:         >80% regular, low sparsity, no clear vocab signal
      - numerical_record:      mostly INT, few STRING, no vocab match

    Returns (document_type, score).
    """
    words = set(vocabulary.keys())
    scores: dict[str, float] = {}

    for doc_type, profile_words in DOCUMENT_PROFILES.items():
        overlap = words & profile_words
        weighted = sum(vocabulary.get(w, 0) for w in overlap)
        scores[doc_type] = weighted / max(len(profile_words), 1)

    # Also initialize structural types
    scores.setdefault("event_register", 0)
    scores.setdefault("regular_table", 0)
    scores.setdefault("numerical_record", 0)

    string_pct = 100 * string_cords / total_cords if total_cords > 0 else 0

    # --- Vocabulary-based boosts ---

    # Astronomical: CORE = HIGH confidence labels that indicate astronomy
    # chaki (Scorpio) is astro-specific — only appears in astro khipus
    # kaki is shared with agriculture so it's NOT core (but boosts when combined)
    core_astro = {"kama", "mama", "paka", "maqa", "chaki"}
    core_astro_hits = len(words & core_astro)  # unique core words
    core_astro_total = sum(vocabulary.get(w, 0) for w in words & core_astro)  # total occurrences
    has_kaki = "kaki" in words
    has_qaqa = "qaqa" in words
    # Strong boost: 2+ unique core words, or 1 core + kaki, or core total >= 3 + kaki
    if (core_astro_hits >= 2
            or (core_astro_hits >= 1 and has_kaki and string_pct < 25)
            or (core_astro_total >= 3 and has_kaki)):
        scores["astronomical_journal"] *= 2.5
        scores["astronomical_journal"] = max(scores["astronomical_journal"], 3.0)
    # Moderate boost: chaki + kaki + qaqa (all HIGH but shared) without core events
    elif core_astro_hits >= 1 and has_kaki and has_qaqa:
        scores["astronomical_journal"] *= 2.0
        scores["astronomical_journal"] = max(scores["astronomical_journal"], 1.5)

    # Agro-pastoral: kaqa/wama/chaqa are agro-specific (not shared with astro)
    agro_specific = {"kaqa", "wama", "chaqa", "siqa"}
    agro_hits = len(words & agro_specific)
    if agro_hits >= 2:
        scores["agro_pastoral"] *= 2.0
        scores["agro_pastoral"] = max(scores["agro_pastoral"], 1.5)
    # Regular full table + agro vocab = strong agro signal
    if agro_hits >= 1 and cluster_regularity > 80 and sparsity < 15:
        scores["agro_pastoral"] *= 1.5

    # Judicial: action + kinship co-occurrence
    action_words = {"taka", "kata", "maqa", "maka", "llalla"}
    kinship_words = {"kaka", "tata", "mama", "pana"}
    action_hits = len(words & action_words)
    kinship_hits = len(words & kinship_words)
    if action_hits >= 1 and kinship_hits >= 1:
        scores["judicial_proceeding"] *= 2.0
        scores["kinship_lineage"] *= 0.6
        # Strong judicial signal: 2+ violence words + 2+ kinship = almost certainly judicial
        if action_hits >= 2 or kinship_hits >= 2:
            scores["judicial_proceeding"] = max(scores["judicial_proceeding"], 2.5)
            # Suppress astro — mama+kaki in a violent context is NOT astronomy
            scores["astronomical_journal"] *= 0.3

    # Oracle/governance text: wapa (intendant) + pi (who?) + chay (this) + papa (father)
    # These are NOT astronomical — suppress astro when governance vocabulary dominates
    governance_words = {"wapa", "pi", "chay", "papa", "pata", "chapa", "tapa"}
    governance_hits = len(words & governance_words)
    governance_total = sum(vocabulary.get(w, 0) for w in words & governance_words)
    if governance_hits >= 4 or governance_total >= 20:
        scores["ritual_oracle"] *= 2.0
        scores["ritual_oracle"] = max(scores["ritual_oracle"], 2.5)
        scores["astronomical_journal"] *= 0.2  # strong suppression

    # Labor: maki + companions
    if "maki" in words:
        labor_companions = {"kiki", "kaki", "tama", "paki", "kata"}
        if len(words & labor_companions) >= 1:
            scores["labor_tribute"] *= 2.0

    # Cadastral: -qa toponyms
    qa_toponyms = {"qaqa", "kaqa", "taqa", "chiqa", "paqa", "naqa", "siqa", "piqa"}
    qa_hits = len(words & qa_toponyms)
    if qa_hits >= 2:
        scores["cadastral_survey"] *= 1.5

    # --- Structure-based detection ---

    # Event register: very sparse (>50% empty cells)
    if sparsity > 50:
        scores["event_register"] = 2.0
        # Penalize types that expect dense data — UNLESS astro-specific words present
        # Astro-exclusive signals: chaki (Scorpio), or multiple core astro words
        has_astro_exclusive = (
            "chaki" in words  # Scorpio: only in astro khipus
            or core_astro_hits >= 2  # 2+ unique core words
            or core_astro_total >= 3  # 3+ total core occurrences (e.g. maqa x3)
        )
        for t in ("agro_pastoral", "cadastral_survey"):
            scores[t] *= 0.3
        if not has_astro_exclusive:
            scores["astronomical_journal"] *= 0.3

    # Regular table: high regularity + low sparsity + multiple colors
    if cluster_regularity > 80 and sparsity < 15:
        if n_colors >= 3:
            scores["regular_table"] = max(scores["regular_table"], 1.5)
        # Penalize event register
        scores["event_register"] *= 0.3

    # Numerical record: very few STRING cords, no clear vocab
    if string_pct < 3 and total_cords > 50:
        top_score = max(scores.values()) if scores else 0
        if top_score < 1.0:
            scores["numerical_record"] = 1.0

    # --- Pick winner ---

    best_type = max(scores, key=scores.get) if scores else "unknown"
    best_score = scores.get(best_type, 0)

    threshold = 0.5 if total_cords < 50 else 1.0
    return (best_type, best_score) if best_score >= threshold else ("unknown", 0)


# --- Architecture detection --------------------------------------------------

def detect_architecture(cords_df: pd.DataFrame) -> str:
    """
    Detect khipu architecture: 0-SEUL, 1-PAIRE, 2+-ARBRE.

    Based on subsidiary depth distribution.
    """
    if "CORD_LEVEL" not in cords_df.columns:
        return "0-SEUL"
    max_level = cords_df["CORD_LEVEL"].max()
    if max_level <= 1:
        return "0-SEUL"
    l2_count = (cords_df["CORD_LEVEL"] == 2).sum()
    l1_count = (cords_df["CORD_LEVEL"] == 1).sum()
    if l1_count > 0 and l2_count / l1_count > 1.5:
        return "2+-ARBRE"
    if l2_count > 0:
        return "1-PAIRE"
    return "0-SEUL"


# --- Translation result ------------------------------------------------------


@dataclass
class TranslationResult:
    """
    Complete multi-level translation of a khipu.

    Attributes
    ----------
    khipu : KhipuRecord
        Metadata about the source khipu.
    cords : list[CordTranslation]
        Level 1: All cords with their translations.
    clusters : list[ClusterRecord]
        Level 2: Cords grouped into structured records.
    document_type : str
        Level 3: Detected document type.
    architecture : str
        Structural type: '0-SEUL', '1-PAIRE', '2+-ARBRE'.
    vocabulary : Counter
        Word frequency counts for STRING cords.
    stats : dict
        Summary statistics.
    """
    khipu: KhipuRecord
    cords: list[CordTranslation]
    clusters: list[ClusterRecord]
    document_type: str
    document_type_score: float
    architecture: str
    vocabulary: Counter
    stats: dict

    # --- Level 1: Cord-level output ------------------------------------------

    def level1_dataframe(self) -> pd.DataFrame:
        """Level 1 export: one row per cord."""
        rows = [c.to_dict() for c in self.cords]
        # Flatten (remove children for DataFrame)
        for r in rows:
            r.pop("children", None)
        return pd.DataFrame(rows)

    # --- Level 2: Record-level output ----------------------------------------

    def level2_records(self) -> list[dict]:
        """Level 2 export: one record per cluster."""
        return [cl.to_dict() for cl in self.clusters]

    # --- Level 3: Document-level output --------------------------------------

    def level3_document(self, lang: str = "en") -> dict:
        """Level 3 export: full document interpretation."""
        return {
            "khipu_id": self.khipu.investigator_num,
            "provenance": self.khipu.provenance,
            "museum": self.khipu.museum_name,
            "document_type": self.document_type,
            "architecture": self.architecture,
            "stats": self.stats,
            "vocabulary": {
                w: {
                    "count": c,
                    "gloss": self._gloss(w, lang),
                    **({"astro": self._domain_gloss(w, lang)}
                       if self._domain_gloss(w, lang) else {}),
                }
                for w, c in self.vocabulary.most_common()
            },
            "clusters": self.level2_records(),
        }

    def _gloss(self, word: str, lang: str) -> str:
        from khipu_translator.glossary_es import GLOSSARY_ES
        base = ""
        if lang == "es" and word in GLOSSARY_ES:
            base = GLOSSARY_ES[word]
        elif word in GLOSSARY:
            fr, en, _ = GLOSSARY[word]
            base = en if lang == "en" else fr
        else:
            # Try onset normalization for Spanish too
            from khipu_translator.dictionary import normalize_onset
            norm = normalize_onset(word)
            if lang == "es" and norm in GLOSSARY_ES:
                base = GLOSSARY_ES[norm]
            elif norm != word and norm in GLOSSARY:
                fr, en, _ = GLOSSARY[norm]
                base = en if lang == "en" else fr
            else:
                morph = analyze_morphology(word, lang)
                if morph.compound_parts:
                    if lang == "es":
                        # For compounds, gloss each part in Spanish
                        parts = []
                        for p in morph.compound_parts:
                            w = p[0]
                            es = GLOSSARY_ES.get(w, p[2])  # fallback to English
                            parts.append(es)
                        base = " + ".join(parts)
                    elif lang == "fr":
                        base = " + ".join(p[1] for p in morph.compound_parts)
                    else:
                        base = " + ".join(p[2] for p in morph.compound_parts)
                elif morph.is_decomposable and morph.root_gloss_en:
                    if lang == "es":
                        g = GLOSSARY_ES.get(morph.root, morph.root_gloss_en)
                    elif lang == "en":
                        g = morph.root_gloss_en
                    else:
                        g = morph.root_gloss_fr
                    if morph.suffixes:
                        suf_str = "+".join(s[1] for s in morph.suffixes)
                        base = f"{g} [{suf_str}]"
                    else:
                        base = g
        return base

    def _domain_gloss(self, word: str, lang: str) -> str:
        """Return domain-specific meaning if available for this document type."""
        if lang == "es":
            from khipu_translator.glossary_es import DOMAIN_GLOSSARIES_ES
            glossary = DOMAIN_GLOSSARIES_ES.get(self.document_type)
            if glossary and word in glossary:
                return glossary[word]
            return ""
        glossary = DOMAIN_GLOSSARIES.get(self.document_type)
        if glossary and word in glossary:
            fr, en = glossary[word]
            return en if lang == "en" else fr
        return ""

    # --- Export methods -------------------------------------------------------

    @staticmethod
    def _ensure_dir(path: str) -> None:
        """Create parent directories if they don't exist."""
        from pathlib import Path
        Path(path).parent.mkdir(parents=True, exist_ok=True)

    def to_json(self, path: str, level: int = 3, lang: str = "en") -> None:
        """Export translation to JSON file."""
        import json
        self._ensure_dir(path)
        if level == 1:
            data = [c.to_dict() for c in self.cords]
        elif level == 2:
            data = self.level2_records()
        else:
            data = self.level3_document(lang=lang)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def to_csv(self, path: str) -> None:
        """Export Level 1 (cord-level) to CSV."""
        self._ensure_dir(path)
        self.level1_dataframe().to_csv(path, index=False, encoding="utf-8")

    def to_xml(self, path: str, lang: str = "en") -> None:
        """Export Level 2 (record-level) to XML."""
        self._ensure_dir(path)
        import xml.etree.ElementTree as ET
        from xml.dom.minidom import parseString

        root = ET.Element("khipu")
        root.set("id", self.khipu.investigator_num)
        root.set("provenance", self.khipu.provenance or "unknown")
        root.set("document_type", self.document_type)
        root.set("architecture", self.architecture)

        for cl in self.clusters:
            cl_elem = ET.SubElement(root, "cluster")
            cl_elem.set("index", str(cl.cluster_index))
            cl_elem.set("total_value", str(cl.total_value))

            for c in cl.cords:
                cord_elem = ET.SubElement(cl_elem, "cord")
                cord_elem.set("id", str(c.cord_id))
                cord_elem.set("type", c.cord_type)
                cord_elem.set("color", c.color)
                cord_elem.set("color_role", c.color_role)
                cord_elem.set("level", str(c.level))

                if c.locke_value is not None:
                    val_elem = ET.SubElement(cord_elem, "value")
                    val_elem.set("confidence", c.locke_confidence)
                    val_elem.text = str(c.locke_value)

                if c.alba_reading:
                    text_elem = ET.SubElement(cord_elem, "reading")
                    text_elem.set("confirmed", str(c.alba_confirmed).lower())
                    text_elem.set("confidence", c.alba_confidence)
                    text_elem.text = c.alba_reading
                    gloss = c.alba_gloss_en if lang == "en" else c.alba_gloss_fr
                    if gloss:
                        gloss_elem = ET.SubElement(cord_elem, "gloss")
                        gloss_elem.text = gloss
                    if c.alba_compound:
                        compound_elem = ET.SubElement(cord_elem, "compound")
                        compound_elem.text = c.alba_compound

                if c.s_prefix:
                    prefix_elem = ET.SubElement(cord_elem, "numeric_prefix")
                    prefix_elem.text = str(c.s_prefix)

                if c.is_l1_null:
                    cord_elem.set("l1_null", "true")

        xml_str = ET.tostring(root, encoding="unicode")
        pretty = parseString(xml_str).toprettyxml(indent="  ")
        with open(path, "w", encoding="utf-8") as f:
            f.write(pretty)

    def to_xlsx(self, path: str, lang: str = "en") -> None:
        """
        Export a human-friendly Excel workbook.

        Sheet 1 "Resume": khipu metadata, statistics, vocabulary with glosses.
        Sheet 2 "Traduction": one row per cord, readable column names, color coding.
        Sheet 3 "Tableur" (if applicable): spreadsheet view for regular-structure
                 khipus (astronomical journals, paired registers).

        Requires openpyxl: pip install khipu-translator[xlsx]
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export.\n"
                "Install it with: pip install khipu-translator[xlsx]"
            )

        self._ensure_dir(path)
        wb = Workbook()

        # --- Color constants ---
        HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
        HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
        TITLE_FONT = Font(bold=True, size=14)
        SUBTITLE_FONT = Font(bold=True, size=11)
        LABEL_FONT = Font(bold=True, size=10)
        CONFIRMED_FONT = Font(bold=True, color="006400")
        UNCONFIRMED_FONT = Font(italic=True, color="8B4513")
        STRING_FILL = PatternFill("solid", fgColor="FFF8DC")
        INT_FILL = PatternFill("solid", fgColor="F0F8FF")
        EMPTY_FILL = PatternFill("solid", fgColor="F5F5F5")
        CLUSTER_FILL = PatternFill("solid", fgColor="E8E8E8")
        THIN_BORDER = Border(
            bottom=Side(style="thin", color="CCCCCC"),
        )
        CLUSTER_BORDER = Border(
            top=Side(style="medium", color="2C3E50"),
        )

        gloss_key = "fr" if lang == "fr" else "en"
        gl = lambda w: self._gloss(w, lang)  # noqa: E731
        dg = lambda w: self._domain_gloss(w, lang)  # noqa: E731

        # =====================================================================
        # Sheet 1: Resume
        # =====================================================================
        ws1 = wb.active
        ws1.title = "Resume"
        ws1.sheet_properties.tabColor = "2C3E50"

        # Title
        ws1["A1"] = f"Khipu {self.khipu.investigator_num}"
        ws1["A1"].font = TITLE_FONT
        ws1.merge_cells("A1:D1")

        # Metadata
        meta = [
            ("Provenance", self.khipu.provenance or "Inconnue"),
            ("Musee", self.khipu.museum_name or "Inconnu"),
            ("Type de document", self.document_type.replace("_", " ").title()),
            ("Architecture", self.architecture),
            ("Cordes totales", self.stats["total_cords"]),
            ("Cordes INT (nombres)", self.stats["int_cords"]),
            ("Cordes STRING (texte)", self.stats["string_cords"]),
            ("Cordes vides", self.stats["empty_cords"]),
            ("Clusters", self.stats["num_clusters"]),
            ("Hits dictionnaire", f"{self.stats['dict_hits']}/{self.stats['string_cords']}"
                f" ({self.stats['coverage_pct']:.0f}%)"),
            ("Valeur numerique totale", f"{self.stats['total_value']:,}"),
        ]
        if self.stats.get("l1_null_cords", 0) > 0:
            meta.append(("Cordes illisibles (L-1)", self.stats["l1_null_cords"]))

        for i, (label, value) in enumerate(meta, start=3):
            ws1.cell(row=i, column=1, value=label).font = LABEL_FONT
            ws1.cell(row=i, column=2, value=str(value))

        # Colors
        row = len(meta) + 4
        ws1.cell(row=row, column=1, value="Couleurs").font = SUBTITLE_FONT
        row += 1
        if self.stats.get("color_distribution"):
            for color, count in sorted(
                self.stats["color_distribution"].items(), key=lambda x: -x[1]
            ):
                role = COLOR_ROLES.get(color.strip(), "")
                ws1.cell(row=row, column=1, value=color.strip())
                ws1.cell(row=row, column=2, value=count)
                ws1.cell(row=row, column=3, value=role)
                row += 1

        # Vocabulary
        row += 1
        ws1.cell(row=row, column=1, value="Vocabulaire").font = SUBTITLE_FONT
        row += 1
        vocab_headers = ["Mot", "Occurrences", "Traduction", "Domaine", "Confirme"]
        for j, h in enumerate(vocab_headers, start=1):
            c = ws1.cell(row=row, column=j, value=h)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
        row += 1
        for word, count in self.vocabulary.most_common():
            ws1.cell(row=row, column=1, value=word).font = (
                CONFIRMED_FONT if word in DICTIONARY else UNCONFIRMED_FONT
            )
            ws1.cell(row=row, column=2, value=count)
            ws1.cell(row=row, column=3, value=gl(word))
            domain = dg(word)
            if domain:
                ws1.cell(row=row, column=4, value=domain)
            ws1.cell(row=row, column=5, value="oui" if word in DICTIONARY else "non")
            row += 1

        # Column widths
        ws1.column_dimensions["A"].width = 25
        ws1.column_dimensions["B"].width = 18
        ws1.column_dimensions["C"].width = 35
        ws1.column_dimensions["D"].width = 35
        ws1.column_dimensions["E"].width = 10

        # Knowledge section (what we know about this khipu)
        from khipu_translator.knowledge import get_knowledge
        knowledge = get_knowledge(self.khipu.investigator_num)
        if knowledge:
            row += 1
            ws1.cell(row=row, column=1, value="What We Know").font = SUBTITLE_FONT
            row += 1
            ws1.cell(row=row, column=1,
                     value=knowledge.get("summary", "")).font = Font(
                         bold=True, size=10)
            row += 1
            ws1.cell(row=row, column=1,
                     value=f"Confidence: {knowledge.get('confidence', '?')}").font = (
                         Font(italic=True, size=10))
            row += 1
            for line in knowledge.get("interpretation", "").strip().split("\n"):
                ws1.cell(row=row, column=1, value=line)
                if line.strip() and not line.startswith(" "):
                    ws1.cell(row=row, column=1).font = Font(bold=True, size=10)
                row += 1
            if knowledge.get("reconstructed_xlsx"):
                row += 1
                ws1.cell(row=row, column=1,
                         value=f"Reconstructed library file: "
                         f"{knowledge['reconstructed_xlsx']}").font = Font(
                             bold=True, color="2E86C1", size=10)
                row += 1
            if knowledge.get("references"):
                row += 1
                ws1.cell(row=row, column=1, value="References:").font = Font(
                    bold=True, size=10)
                row += 1
                for ref in knowledge["references"]:
                    ws1.cell(row=row, column=1, value=f"  - {ref}")
                    row += 1

        # =====================================================================
        # Sheet 2: Traduction
        # =====================================================================
        ws2 = wb.create_sheet("Traduction")
        ws2.sheet_properties.tabColor = "E67E22"

        headers = [
            "Cluster", "Position", "Niveau", "Couleur", "Role couleur",
            "Type", "Noeuds", "Valeur", "Lecture", "Traduction",
            "Domaine", "Confiance", "Prefixe num.",
        ]
        for j, h in enumerate(headers, start=1):
            c = ws2.cell(row=1, column=j, value=h)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal="center")
        ws2.freeze_panes = "A2"

        L2_FILL = PatternFill("solid", fgColor="EBF5FB")
        L3_FILL = PatternFill("solid", fgColor="D5F5E3")

        row = 2
        prev_cluster = None
        # Track L2/L3 row ranges for grouping
        group_ranges: list[tuple[int, int, int]] = []  # (start, end, level)

        for cl in self.clusters:
            # Cluster separator row
            if prev_cluster is not None:
                for j in range(1, len(headers) + 1):
                    c = ws2.cell(row=row, column=j, value="")
                    c.border = CLUSTER_BORDER
                row += 1

            def _write_cord(
                cord: CordTranslation, cluster_idx: int, r: int,
            ) -> int:
                indent = "    " * (cord.level - 1)

                ws2.cell(row=r, column=1, value=cluster_idx)
                ws2.cell(row=r, column=2, value=cord.ordinal)
                ws2.cell(row=r, column=3, value=f"{indent}L{cord.level}")
                ws2.cell(row=r, column=4, value=cord.color.strip())
                ws2.cell(row=r, column=5, value=cord.color_role)

                # Choose fill based on level
                if cord.level >= 3:
                    level_fill = L3_FILL
                elif cord.level == 2:
                    level_fill = L2_FILL
                else:
                    level_fill = None

                # Column 7: knot sequence (always written)
                if cord.knot_sequence:
                    ws2.cell(row=r, column=7, value=cord.knot_sequence)

                if cord.cord_type == "STRING":
                    ws2.cell(row=r, column=6, value="TEXTE")
                    reading = cord.alba_reading or "(illisible)"
                    c_read = ws2.cell(row=r, column=9, value=reading)
                    if cord.alba_confirmed:
                        c_read.font = CONFIRMED_FONT
                    else:
                        c_read.font = UNCONFIRMED_FONT
                    gloss_val = (cord.alba_gloss_fr if lang == "fr"
                                 else cord.alba_gloss_en) or ""
                    ws2.cell(row=r, column=10, value=gloss_val)
                    domain_val = dg(cord.alba_reading) if cord.alba_reading else ""
                    if domain_val:
                        ws2.cell(row=r, column=11, value=domain_val)
                    ws2.cell(row=r, column=12, value=cord.alba_confidence)
                    if cord.s_prefix:
                        ws2.cell(row=r, column=13, value=cord.s_prefix)
                    base_fill = STRING_FILL if cord.level == 1 else level_fill
                    for j in range(1, len(headers) + 1):
                        ws2.cell(row=r, column=j).fill = base_fill or STRING_FILL

                elif cord.cord_type == "INT":
                    ws2.cell(row=r, column=6, value="NOMBRE")
                    if cord.locke_value is not None:
                        ws2.cell(row=r, column=8, value=int(cord.locke_value))
                    ws2.cell(row=r, column=12, value=cord.locke_confidence)
                    base_fill = INT_FILL if cord.level == 1 else level_fill
                    for j in range(1, len(headers) + 1):
                        ws2.cell(row=r, column=j).fill = base_fill or INT_FILL

                else:
                    ws2.cell(row=r, column=6, value="VIDE")
                    base_fill = EMPTY_FILL if cord.level == 1 else level_fill
                    for j in range(1, len(headers) + 1):
                        ws2.cell(row=r, column=j).fill = base_fill or EMPTY_FILL

                for j in range(1, len(headers) + 1):
                    ws2.cell(row=r, column=j).border = THIN_BORDER

                r += 1

                # Write children and track their row range for grouping
                if cord.children:
                    child_start = r
                    for child in cord.children:
                        r = _write_cord(child, cluster_idx, r)
                    child_end = r - 1
                    outline_level = min(cord.level, 2)  # L1→group level 1, L2→level 2
                    group_ranges.append((child_start, child_end, outline_level))

                return r

            for cord in cl.cords:
                row = _write_cord(cord, cl.cluster_index, row)

            prev_cluster = cl.cluster_index

        # Apply row grouping (collapsible outlines)
        for start, end, level in group_ranges:
            if start <= end:
                ws2.row_dimensions.group(start, end,
                                         outline_level=level, hidden=False)

        # Column widths (13 columns now)
        # Cluster, Position, Niveau, Couleur, Role, Type, Noeuds, Valeur, Lecture, Traduction, Domaine, Confiance, Prefixe
        widths = [8, 8, 8, 8, 18, 8, 18, 10, 20, 30, 30, 12, 10]
        for j, w in enumerate(widths, start=1):
            col_letter = chr(64 + j) if j <= 26 else "Z"
            ws2.column_dimensions[col_letter].width = w

        wb.save(path)


    # --- Summary -------------------------------------------------------------

    def summary(self, lang: str = "en") -> str:
        """Human-readable summary of the translation."""
        s = self.stats
        lines = [
            f"{'=' * 60}",
            f"  {self.khipu.investigator_num}",
            f"{'=' * 60}",
            f"  Provenance: {self.khipu.provenance or 'Unknown'}",
            f"  Museum: {self.khipu.museum_name or 'Unknown'}",
            f"  Cords: {s['total_cords']} (INT: {s['int_cords']}, "
            f"STRING: {s['string_cords']}, EMPTY: {s['empty_cords']})",
            f"  Clusters: {s['num_clusters']}",
            f"  Architecture: {self.architecture}",
        ]

        # Schema detection
        from khipu_translator.schema import detect_schema
        schema = detect_schema(self)
        lines.append(f"  Schema: {schema.schema_type.upper()} (depth L{schema.max_depth})")

        # Date detection
        from khipu_translator.dating import extract_date
        date = extract_date(self)
        if date:
            day_str = f"/{date.day:02d}" if date.day else ""
            if date.year_ce and date.month:
                lines.append(
                    f"  Date: {date.year_ce}/{date.month:02d}{day_str} CE"
                    f" ({date.month_name or ''}, {date.gregorian_month or ''})"
                    f" [Mode {date.mode}]"
                )
            elif date.month:
                lines.append(
                    f"  Month: {date.month} ({date.month_name or ''}, "
                    f"{date.gregorian_month or ''}) [Mode {date.mode}]"
                )

        if s.get("l1_null_cords", 0) > 0:
            lines.append(
                f"  L-1 (unreadable): {s['l1_null_cords']} cords "
                f"(NULL turns in OKR)"
            )

        # Colors
        if s.get("color_distribution"):
            top_colors = sorted(
                s["color_distribution"].items(), key=lambda x: -x[1]
            )[:5]
            color_str = ", ".join(f"{c}({n})" for c, n in top_colors)
            lines.append(f"  Colors: {color_str}")

        if s["string_cords"] > 0:
            lines.append("")
            lines.append(
                f"  Dictionary hits: {s['dict_hits']}/{s['string_cords']} "
                f"({s['coverage_pct']:.0f}%)"
            )
            lines.append(f"  Document type: {self.document_type}")
            lines.append(f"  Vocabulary ({len(self.vocabulary)} unique words):")
            for word, count in self.vocabulary.most_common(20):
                gloss = self._gloss(word, lang)
                from khipu_translator.dictionary import normalize_onset
                is_known = (word in DICTIONARY
                            or normalize_onset(word) in DICTIONARY)
                morph = analyze_morphology(word)
                # * = dictionary match, + = decomposable (morphology/compound), ~ = unknown
                if is_known:
                    marker = "*"
                elif morph.is_decomposable:
                    marker = "+"
                else:
                    marker = "~"
                compound = ""
                if morph.compound_parts:
                    parts = " + ".join(
                        p[2] if lang == "en" else p[1]
                        for p in morph.compound_parts
                    )
                    compound = f" -> {parts}"
                astro = self._domain_gloss(word, lang)
                astro_str = f"  ({astro})" if astro else ""
                lines.append(
                    f"    {marker}{word:20s} x{count:<3d} {gloss}{compound}{astro_str}"
                )
        else:
            lines.append("")
            lines.append("  This khipu is purely numerical (no STRING cords).")
            if s.get("total_value"):
                lines.append(f"  Total numerical value: {s['total_value']:,}")

        lines.append(f"\n{'=' * 60}")
        return "\n".join(lines)


# --- Main translate function -------------------------------------------------


def _split_knots_by_position(terminal_knots: pd.DataFrame) -> list[list[dict]]:
    """
    Split a cord's terminal knots into word groups using knot_value_type
    positional resets (Locke Word Splitting).

    When knot_value_type INCREASES from one knot to the next, the scribe
    has moved back UP the cord = start of a new word.

    Only applies to cords with 3+ terminal knots. Cords with 2 knots
    = always 1 word.

    Returns list of knot groups, each group = one word.
    """
    knots = []
    for _, knot in terminal_knots.iterrows():
        tc = str(knot.get("TYPE_CODE", "")).strip()
        turns = knot.get("NUM_TURNS")
        kvt = knot.get("knot_value_type", 0) or 0
        knots.append({"tc": tc, "turns": turns, "kvt": int(kvt)})

    # Split applies to ALL cords (including 2-knot).
    # 19% of 2-L/E cords have ascending kvt = real word boundary.
    # Single-syllable results (pa, ta, ma, ka, qa, y, na, pi) are
    # real Quechua grammatical morphemes, not artifacts.
    if len(knots) < 2:
        return [knots]

    # Check if kvt data is usable (not all zeros or all same value)
    kvt_vals = [k["kvt"] for k in knots]
    if len(set(kvt_vals)) <= 1:
        return [knots]  # no positional info → treat as single word

    # Split on kvt increases
    words = []
    current_word = [knots[0]]
    prev_kvt = knots[0]["kvt"]

    for knot in knots[1:]:
        kvt = knot["kvt"]
        if kvt > prev_kvt and current_word:
            words.append(current_word)
            current_word = [knot]
        else:
            current_word.append(knot)
        prev_kvt = kvt

    if current_word:
        words.append(current_word)

    return words


def _knots_to_word(knot_group: list[dict]) -> Optional[str]:
    """
    Convert a group of knots into a word using syllabary v3.

    Onset polyphony applies to the FIRST knot of this word group,
    coda applies to the LAST knot.
    """
    syllables = []
    n = len(knot_group)

    for i, knot in enumerate(knot_group):
        tc = knot["tc"]
        turns = knot["turns"]
        is_first = (i == 0)

        if tc == "E":
            syllables.append(FIGURE_EIGHT_SYLLABLE)
        elif tc == "L":
            if pd.isna(turns) or turns is None:
                return None
            turns = int(turns)
            if is_first:
                syl = TURNS_TO_ONSET.get(turns)
            else:
                syl = TURNS_TO_SYLLABLE.get(turns)
            if syl is None:
                return None
            syllables.append(syl)

    return "".join(syllables) if syllables else None


_AFFIX_LABELS = {"GEN", "ACC", "TOP", "INF", "POSS", "OBL", "REFL",
                  "DIR", "PASS", "RECIP", "CAUS", "LIM", "INTER",
                  "1OBJ", "DIM"}


def _strip_affixes(word: str) -> str:
    """Strip grammatical prefix/suffix labels from a word to get the bare root.

    Handles multiple affixes: 'tama-1OBJ-DIM' → 'tama',
    'GEN-chayyy' → 'chayyy', 'ACC-papa-GEN' → 'papa'.
    """
    if "-" not in word:
        return word
    parts = word.split("-")
    # Keep only parts that are NOT affix labels
    core = [p for p in parts if p not in _AFFIX_LABELS]
    return "".join(core) if core else word


def _read_alba_word(terminal_knots: pd.DataFrame) -> tuple[Optional[str], bool]:
    """
    Decode a STRING cord's terminal knots into proposed Quechua word(s).

    Uses Locke Word Splitting: knot_value_type positional resets split
    a single cord into multiple words. Then onset polyphony (v3) is
    applied per WORD, not per cord.

    Returns
    -------
    (word, is_l1_null) : tuple
        word = single string (words joined with ' ') or None.
        is_l1_null flags NULL turns in OKR.
    """
    # Check for NULL turns first
    for _, knot in terminal_knots.iterrows():
        tc = str(knot.get("TYPE_CODE", "")).strip()
        if tc == "L":
            turns = knot.get("NUM_TURNS")
            if pd.isna(turns) or turns is None:
                return None, True

    # Split by positional resets
    word_groups = _split_knots_by_position(terminal_knots)

    # Convert each group to a word
    words = []
    for group in word_groups:
        w = _knots_to_word(group)
        if w:
            words.append(w)

    if not words:
        return None, False

    # If only one word, return it directly
    if len(words) == 1:
        return words[0], False

    # Post-processing: attach isolated suffixes to adjacent root words.
    # Quechua suffixes (pa=GEN, ta=ACC, ka=PASS, etc.) that appear as
    # isolated fragments after splitting should be attached to the
    # nearest root word with a hyphen: "papa-GEN" not "papa pa".
    from khipu_translator.dictionary import QUECHUA_SUFFIXES
    suffix_set = set(QUECHUA_SUFFIXES.keys())

    attached = []
    i = 0
    while i < len(words):
        w = words[i]
        if w in suffix_set:
            label = QUECHUA_SUFFIXES[w][0]
            if attached:
                # Suffix after a root → attach to previous word
                attached[-1] = f"{attached[-1]}-{label}"
            else:
                # Suffix at start → attach to next word as prefix
                if i + 1 < len(words):
                    words[i + 1] = f"{label}-{words[i + 1]}"
                else:
                    attached.append(f"-{label}")
        else:
            attached.append(w)
        i += 1

    return " ".join(attached), False


def translate(
    khipu_name: str,
    db: Optional[KhipuDB] = None,
    lang: str = "en",
) -> TranslationResult:
    """
    Translate a khipu from the OKR database.

    Parameters
    ----------
    khipu_name : str
        Khipu identifier (e.g. 'UR039', 'AS030', 'UR268').
    db : KhipuDB, optional
        Database connection. If None, creates one (auto-downloads OKR).
    lang : str
        Language for glosses: 'en' (English) or 'fr' (French).

    Returns
    -------
    TranslationResult
        Multi-level translation with export capabilities.

    Examples
    --------
    >>> result = translate("UR039")
    >>> print(result.summary())
    >>> result.to_json("UR039_level3.json", level=3)
    >>> result.to_csv("UR039_cords.csv")
    >>> result.to_xml("UR039.xml")
    """
    close_db = False
    if db is None:
        db = KhipuDB()
        close_db = True

    try:
        khipu = db.get_khipu(khipu_name)
        cords_df = db.get_cords(khipu.khipu_id)
        knots_df = db.get_knots(khipu.khipu_id)
    finally:
        if close_db:
            db.close()

    # --- Process each cord ---------------------------------------------------

    cord_translations: dict[int, CordTranslation] = {}
    vocabulary: Counter = Counter()
    l1_null_count = 0
    color_distribution: Counter = Counter()

    for _, cord_row in cords_df.iterrows():
        cid = int(cord_row["CORD_ID"])
        cord_knots = knots_df[knots_df["CORD_ID"] == cid].sort_values(
            ["CLUSTER_ORDINAL", "KNOT_ORDINAL"]
        )

        # Classify cord
        terminal = cord_knots[cord_knots["TYPE_CODE"].isin(["L", "E"])]
        simple = cord_knots[cord_knots["TYPE_CODE"] == "S"]
        ctype = "EMPTY"
        if len(cord_knots) > 0:
            ctype = "STRING" if len(terminal) >= 2 else "INT"

        # Color
        color = str(cord_row.get("color", "?"))
        color_distribution[color] += 1
        color_role = COLOR_ROLES.get(color, "unknown")

        # Build knot sequence string
        knot_seq_parts = []
        for _, kn in cord_knots.iterrows():
            tc = str(kn.get("TYPE_CODE", "")).strip()
            turns = kn.get("NUM_TURNS")
            kvt = kn.get("knot_value_type", 0) or 0
            if tc == "S":
                knot_seq_parts.append(f"S{int(kvt)}")
            elif tc == "L":
                t = int(turns) if pd.notna(turns) else "?"
                knot_seq_parts.append(f"L{t}")
            elif tc == "E":
                knot_seq_parts.append("E")
        knot_sequence = " ".join(knot_seq_parts)

        # Decode channels
        locke_val = None
        locke_conf = "none"
        alba_word = None
        alba_gloss_en = None
        alba_gloss_fr = None
        alba_confirmed = False
        alba_confidence = "none"
        alba_compound = None
        s_prefix = 0
        is_l1_null = False

        if ctype == "INT":
            # Numerical channel
            knot_dicts = cord_knots.to_dict("records")
            lv = decode_locke_value(knot_dicts)
            if lv:
                locke_val = lv.value
                locke_conf = lv.confidence

        elif ctype == "STRING":
            # Textual channel — with Locke Word Splitting
            alba_word, is_l1_null = _read_alba_word(terminal)
            if is_l1_null:
                l1_null_count += 1

            if alba_word:
                from khipu_translator.dictionary import normalize_onset

                # Multi-word: split on spaces and check each word
                split_words = alba_word.split()
                if len(split_words) > 1:
                    # Multiple words from positional splitting
                    all_confirmed = True
                    glosses_en = []
                    glosses_fr = []
                    for sw in split_words:
                        # Strip affix labels for dictionary lookup
                        bare = _strip_affixes(sw)
                        bare_ok = (bare in DICTIONARY
                                   or normalize_onset(bare) in DICTIONARY)
                        if not bare_ok:
                            bare_morph = analyze_morphology(bare, lang=lang)
                            if not bare_morph.is_decomposable:
                                all_confirmed = False
                        vocabulary[bare] += 1
                        # Get gloss for display (use bare root)
                        sw_morph = analyze_morphology(bare, lang=lang)
                        glosses_en.append(sw_morph.root_gloss_en or sw)
                        glosses_fr.append(sw_morph.root_gloss_fr or sw)

                    alba_confirmed = all_confirmed
                    alba_gloss_en = " + ".join(glosses_en)
                    alba_gloss_fr = " + ".join(glosses_fr)
                    alba_confidence = "high" if all_confirmed else "medium"
                    alba_compound = " + ".join(split_words)
                else:
                    # Single word (no split or 2-knot cord)
                    bare = _strip_affixes(alba_word)
                    alba_confirmed = (
                        bare in DICTIONARY
                        or normalize_onset(bare) in DICTIONARY
                    )
                    morph = analyze_morphology(bare, lang=lang)
                    alba_gloss_en = morph.root_gloss_en or None
                    alba_gloss_fr = morph.root_gloss_fr or None

                    if morph.compound_parts:
                        alba_compound = " + ".join(
                            f"{p[0]}({p[2]})" for p in morph.compound_parts
                        )
                        alba_confidence = "medium"
                    elif alba_confirmed:
                        alba_confidence = "high"
                    elif morph.is_decomposable:
                        alba_confidence = "medium"
                    else:
                        alba_confidence = "low"

                # Count in vocabulary (multi-word already counted above)
                if " " not in alba_word:
                    vocabulary[_strip_affixes(alba_word)] += 1

            # Numerical prefix (S-knots on STRING cords)
            if len(simple) > 0:
                s_prefix = int(simple["knot_value_type"].sum())

        # Level
        cord_level = int(cord_row.get("CORD_LEVEL", 1))

        ct = CordTranslation(
            cord_id=cid,
            cord_type=ctype,
            level=cord_level,
            ordinal=int(cord_row["CLUSTER_ORDINAL"]) if pd.notna(
                cord_row.get("CLUSTER_ORDINAL")) else 0,
            global_ordinal=int(cord_row["CORD_ORDINAL"]) if pd.notna(
                cord_row.get("CORD_ORDINAL")) else 0,
            color=color,
            color_role=color_role,
            locke_value=locke_val,
            locke_confidence=locke_conf,
            alba_reading=alba_word,
            alba_gloss_en=alba_gloss_en,
            alba_gloss_fr=alba_gloss_fr,
            alba_confirmed=alba_confirmed,
            alba_confidence=alba_confidence,
            alba_compound=alba_compound,
            knot_sequence=knot_sequence,
            s_prefix=s_prefix,
            is_l1_null=is_l1_null,
            parent_cord_id=(
                int(cord_row["PENDANT_FROM"])
                if pd.notna(cord_row.get("PENDANT_FROM"))
                else None
            ),
        )
        cord_translations[cid] = ct

    # --- Build tree ----------------------------------------------------------

    for ct in cord_translations.values():
        if ct.parent_cord_id and ct.parent_cord_id in cord_translations:
            cord_translations[ct.parent_cord_id].children.append(ct)

    for ct in cord_translations.values():
        ct.children.sort(key=lambda c: c.global_ordinal)

    # --- Group into clusters (Level 2) ---------------------------------------

    clusters_map: dict = defaultdict(list)
    for _, cord_row in cords_df.iterrows():
        cid = int(cord_row["CORD_ID"])
        level = int(cord_row.get("CORD_LEVEL", 1))
        if level == 1:  # only L1 cords form cluster rows
            cl_id = cord_row.get("CLUSTER_ID")
            key = cl_id if pd.notna(cl_id) else f"_orphan_{cid}"
            clusters_map[key].append(cord_translations[cid])

    for key in clusters_map:
        clusters_map[key].sort(key=lambda c: c.global_ordinal)

    sorted_keys = sorted(
        clusters_map.keys(),
        key=lambda k: min(c.global_ordinal for c in clusters_map[k]),
    )

    clusters = [
        ClusterRecord(
            cluster_index=i + 1,
            cluster_id=k if not str(k).startswith("_orphan") else None,
            cords=clusters_map[k],
        )
        for i, k in enumerate(sorted_keys)
    ]

    # --- Architecture --------------------------------------------------------

    architecture = detect_architecture(cords_df)

    # --- Document type detection (Level 3) -----------------------------------

    # --- Statistics (compute before doc type, needed for structural signals) --

    all_cords = list(cord_translations.values())
    string_cords = [c for c in all_cords if c.cord_type == "STRING" and c.alba_reading]
    dict_hits = sum(1 for c in string_cords if c.alba_confirmed)

    n_string = len(string_cords)
    stats = {
        "total_cords": len(all_cords),
        "int_cords": sum(1 for c in all_cords if c.cord_type == "INT"),
        "string_cords": n_string,
        "empty_cords": sum(1 for c in all_cords if c.cord_type == "EMPTY"),
        "l1_null_cords": l1_null_count,
        "dict_hits": dict_hits,
        "coverage_pct": (100 * dict_hits / n_string) if n_string else 0,
        "num_clusters": len(clusters),
        "unique_words": len(vocabulary),
        "total_value": sum(c.locke_value or 0 for c in all_cords if c.cord_type == "INT"),
        "color_distribution": dict(color_distribution),
    }

    # --- Document type detection (Level 3) — uses structural stats -----------

    # Compute structural signals for classification
    _data_cl = [cl for cl in clusters if len(cl.cords) > 1
                and any(c.cord_type != "EMPTY" for c in cl.cords)]
    _cl_sizes = [len(cl.cords) for cl in _data_cl]
    _cluster_regularity = 0.0
    if _cl_sizes:
        _size_counts = Counter(_cl_sizes)
        _mode_n = _size_counts.most_common(1)[0][1]
        _cluster_regularity = 100 * _mode_n / len(_cl_sizes)
    _total_cells = sum(len(cl.cords) for cl in _data_cl)
    _empty_cells = sum(1 for cl in _data_cl for c in cl.cords if c.cord_type == "EMPTY")
    _sparsity = 100 * _empty_cells / _total_cells if _total_cells else 0
    _n_colors = len([c for c, n in color_distribution.items()
                     if n > 5 and c.strip() != "?"])

    doc_type, doc_score = detect_document_type(
        vocabulary,
        color_counts=dict(color_distribution),
        total_cords=len(all_cords),
        string_cords=n_string,
        num_clusters=len(clusters),
        cluster_regularity=_cluster_regularity,
        sparsity=_sparsity,
        n_colors=_n_colors,
    )

    result = TranslationResult(
        khipu=khipu,
        cords=all_cords,
        clusters=clusters,
        document_type=doc_type,
        document_type_score=doc_score,
        architecture=architecture,
        vocabulary=vocabulary,
        stats=stats,
    )

    # Post-processing: if a date is detected, remove date-zone words
    # from vocabulary. These "words" are actually numbers (dates) read
    # through the syllabary — not real textual content.
    from khipu_translator.dating import extract_date
    date = extract_date(result)
    if date and date.mode in ("A", "AB"):
        # The date zone = first 3 L1 cords (year, month, day)
        # Find their cord IDs and remove their readings from vocabulary
        date_zone_count = 0
        max_date_cords = 3  # year + month + day
        for cl in clusters:
            for c in cl.cords:
                if c.level == 1 and date_zone_count < max_date_cords:
                    if c.alba_reading and c.alba_reading in vocabulary:
                        vocabulary[c.alba_reading] -= 1
                        if vocabulary[c.alba_reading] <= 0:
                            del vocabulary[c.alba_reading]
                    date_zone_count += 1
            if date_zone_count >= max_date_cords:
                break

    return result
